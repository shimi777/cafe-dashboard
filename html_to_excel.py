"""
HTML to Excel Parser Module - מודול לעיבוד קובץ HTML של דוחות פעולות ממערכת POS
גרסה מתוקנת עם התאמה למבנה ה-HTML הספציפי
"""

import re
from datetime import datetime
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io


def parse_html_transactions(html_content):
    """
    Parse transaction details from HTML
    Returns list of transactions with their details

    מבנה ה-HTML:
    - כל טרנזקציה בתוך div.data-block
    - Header עם פרטי העסקה בתוך div.trans-header
    - פריטים בתוך div.table-contents
    - תשלומים בתוך div.table-tenders
    - סיכום בתוך div.table-totals
    """
    soup = BeautifulSoup(html_content, 'html.parser')

    transactions = []

    # Find all data blocks (each block = one transaction)
    data_blocks = soup.find_all('div', class_='data-block')

    for block in data_blocks:
        try:
            transaction = parse_single_block(block)
            if transaction:
                transactions.append(transaction)
        except Exception as e:
            print(f"Error parsing block: {e}")
            continue

    return transactions


def parse_single_block(block):
    """
    Parse a single transaction block
    """
    # Extract transaction header
    trans_header = block.find('div', class_='trans-header')
    if not trans_header:
        return None

    # Parse header fields by looking at item-title and header-num pairs
    header_data = {}
    text_divs = trans_header.find_all('div', class_='text')

    for text_div in text_divs:
        title_div = text_div.find('div', class_='item-title')
        value_span = text_div.find('span', class_='header-num')

        if title_div and value_span:
            title = title_div.get_text(strip=True)
            value = value_span.get_text(strip=True)
            header_data[title] = value

    # Extract required fields
    order_id = header_data.get('הזמנה', '')
    invoice_num = header_data.get('חשבונית מס', '')
    transaction_type = header_data.get('סוג עסקה', '')
    z_number = header_data.get('זד מספר', '')
    datetime_str = header_data.get('תאריך', '')
    register = header_data.get('קופה', '')
    customer_name = header_data.get('שם לקוח', '')
    customer_code = header_data.get('קוד לקוח', '')

    # Parse date and time
    trans_date = None
    trans_time = None
    if datetime_str:
        try:
            trans_datetime = datetime.strptime(datetime_str, '%d/%m/%Y %H:%M')
            trans_date = trans_datetime.date()
            trans_time = trans_datetime.time()
        except ValueError:
            # Try alternative format
            try:
                trans_datetime = datetime.strptime(datetime_str, '%d/%m/%Y')
                trans_date = trans_datetime.date()
                trans_time = datetime.strptime('00:00', '%H:%M').time()
            except:
                return None

    if not trans_date:
        return None

    # Get items from table-contents
    items = []
    table_contents = block.find('div', class_='table-contents')

    if table_contents:
        # Find all item rows (not in tender or totals sections)
        all_item_divs = table_contents.find_all('div', recursive=False)

        for item_container in all_item_divs:
            item_row = item_container.find('div', class_='item-row')
            if not item_row:
                continue

            # Get item name from span
            item_span = item_row.find('span')
            if not item_span:
                continue

            item_name = item_span.get_text(strip=True)
            if not item_name:
                continue

            # Get item code if available
            item_code_div = item_row.find('div', class_='item-code')
            item_code = ''
            if item_code_div:
                item_code = item_code_div.get_text(strip=True).replace('קוד פריט', '').strip()

            # Get numeric values
            num_elements = item_row.find_all('div', class_='num')

            # Expected order: כמות, מחיר ליחידה, חייב מע"מ, מחיר מכירה, קוד מע"מ, מע"מ
            if len(num_elements) >= 6:
                try:
                    quantity = float(num_elements[0].get_text(strip=True).replace(',', ''))
                    unit_price = float(num_elements[1].get_text(strip=True).replace(',', ''))
                    taxable_amount = float(num_elements[2].get_text(strip=True).replace(',', ''))
                    sale_price = float(num_elements[3].get_text(strip=True).replace(',', ''))
                    # num_elements[4] is VAT code (text like "מעמ")
                    vat_amount = float(num_elements[5].get_text(strip=True).replace(',', ''))

                    # Get cashier name from text div in num-4 wrapper
                    cashier = ''
                    text_div = item_row.find('div', class_='text-2')
                    if text_div:
                        cashier_text = text_div.find('div', class_='text')
                        if cashier_text:
                            cashier = cashier_text.get_text(strip=True)

                    items.append({
                        'name': item_name,
                        'code': item_code,
                        'quantity': quantity,
                        'unit_price': unit_price,
                        'taxable_amount': taxable_amount,
                        'total_price': sale_price,
                        'vat_amount': vat_amount,
                        'cashier': cashier
                    })
                except (ValueError, IndexError) as e:
                    continue

    # Get payment methods from table-tenders
    payments = []
    table_tenders = block.find('div', class_='table-tenders')

    if table_tenders:
        tender_rows = table_tenders.find_all('div', class_='tender-row')

        for tender_row in tender_rows:
            payment_data = {}
            text_divs = tender_row.find_all('div', class_='text')

            for text_div in text_divs:
                title_div = text_div.find('div', class_='item-title')
                value_span = text_div.find('span', class_='tender-num')

                if title_div and value_span:
                    title = title_div.get_text(strip=True)
                    value = value_span.get_text(strip=True)

                    if 'צורת תשלום' in title:
                        payment_data['method'] = value
                    elif 'סכום' in title:
                        try:
                            payment_data['amount'] = float(value.replace(',', ''))
                        except:
                            payment_data['amount'] = 0
                    elif 'מספר אישור' in title:
                        payment_data['approval'] = value
                    elif 'סימוכין' in title:
                        payment_data['reference'] = value

            if payment_data.get('method'):
                payments.append(payment_data)

    # Get totals from table-totals
    total_items = 0
    total_vat = 0
    total_amount = 0

    table_totals = block.find('div', class_='table-totals')
    if table_totals:
        totals_row = table_totals.find('div', class_='totals-row')
        if totals_row:
            tender_nums = totals_row.find_all('span', class_='tender-num')
            if len(tender_nums) >= 3:
                try:
                    total_items = float(tender_nums[0].get_text(strip=True).replace(',', ''))
                    total_vat = float(tender_nums[1].get_text(strip=True).replace(',', ''))
                    total_amount = float(tender_nums[2].get_text(strip=True).replace(',', ''))
                except:
                    pass

    # If totals not found, calculate from items
    if total_amount == 0 and items:
        total_amount = sum(item['total_price'] for item in items)
        total_vat = sum(item['vat_amount'] for item in items)
        total_items = sum(item['taxable_amount'] for item in items)

    # Skip cancelled transactions (total = 0)
    if total_amount == 0:
        return None

    return {
        'order_id': order_id,
        'invoice_num': invoice_num,
        'transaction_type': transaction_type,
        'z_number': z_number,
        'date': trans_date,
        'time': trans_time,
        'register': register,
        'customer_name': customer_name,
        'customer_code': customer_code,
        'items': items,
        'payments': payments,
        'total_items': total_items,
        'total_vat': total_vat,
        'total': total_amount
    }


def create_daily_summary(transactions):
    """
    Summarize transactions by date
    Returns dataframe with daily sales summary
    """
    if not transactions:
        return pd.DataFrame(columns=['date', 'total_sales', 'transaction_count', 'items_count'])

    daily_data = {}

    for trans in transactions:
        trans_date = trans['date']
        total = trans['total']

        if trans_date not in daily_data:
            daily_data[trans_date] = {
                'date': trans_date,
                'total_sales': 0,
                'transaction_count': 0,
                'items_count': 0,
                'total_vat': 0
            }

        daily_data[trans_date]['total_sales'] += total
        daily_data[trans_date]['transaction_count'] += 1
        daily_data[trans_date]['items_count'] += len(trans['items'])
        daily_data[trans_date]['total_vat'] += trans.get('total_vat', 0)

    # Convert to dataframe
    daily_df = pd.DataFrame(list(daily_data.values()))
    if not daily_df.empty:
        daily_df = daily_df.sort_values('date')

    return daily_df


def create_detailed_transactions_df(transactions):
    """
    Create detailed transactions dataframe (one row per transaction)
    """
    if not transactions:
        return pd.DataFrame(columns=['Order ID', 'Invoice', 'Date', 'Time', 'Item Count', 'Total Amount'])

    records = []

    for trans in transactions:
        # Get primary payment method
        payment_method = ''
        if trans.get('payments'):
            # Find the payment with largest positive amount
            positive_payments = [p for p in trans['payments'] if p.get('amount', 0) > 0]
            if positive_payments:
                payment_method = max(positive_payments, key=lambda x: x.get('amount', 0)).get('method', '')

        records.append({
            'Order ID': trans['order_id'],
            'Invoice': trans['invoice_num'],
            'Date': trans['date'],
            'Time': trans['time'],
            'Register': trans.get('register', ''),
            'Payment Method': payment_method,
            'Item Count': len(trans['items']),
            'Total Items': trans.get('total_items', 0),
            'Total VAT': trans.get('total_vat', 0),
            'Total Amount': trans['total']
        })

    return pd.DataFrame(records)


def create_items_summary_df(transactions):
    """
    Create item-level summary (aggregated by item name)
    """
    if not transactions:
        return pd.DataFrame(columns=['item_name', 'quantity', 'total_amount', 'transaction_count'])

    items_data = {}

    for trans in transactions:
        for item in trans['items']:
            item_name = item['name']

            if item_name not in items_data:
                items_data[item_name] = {
                    'item_name': item_name,
                    'item_code': item.get('code', ''),
                    'quantity': 0,
                    'total_amount': 0,
                    'total_vat': 0,
                    'transaction_count': 0,
                    'avg_unit_price': 0
                }

            items_data[item_name]['quantity'] += item['quantity']
            items_data[item_name]['total_amount'] += item['total_price']
            items_data[item_name]['total_vat'] += item.get('vat_amount', 0)
            items_data[item_name]['transaction_count'] += 1

    # Calculate average unit price
    for item_name in items_data:
        if items_data[item_name]['quantity'] > 0:
            items_data[item_name]['avg_unit_price'] = (
                    items_data[item_name]['total_amount'] / items_data[item_name]['quantity']
            )

    items_df = pd.DataFrame(list(items_data.values()))
    if not items_df.empty:
        items_df = items_df.sort_values('total_amount', ascending=False)

    return items_df


def create_items_detail_df(transactions):
    """
    Create item-level detail (one row per item sold)
    """
    if not transactions:
        return pd.DataFrame()

    records = []

    for trans in transactions:
        for item in trans['items']:
            records.append({
                'Date': trans['date'],
                'Time': trans['time'],
                'Order ID': trans['order_id'],
                'Item Name': item['name'],
                'Item Code': item.get('code', ''),
                'Quantity': item['quantity'],
                'Unit Price': item['unit_price'],
                'Total Price': item['total_price'],
                'VAT Amount': item.get('vat_amount', 0),
                'Cashier': item.get('cashier', '')
            })

    return pd.DataFrame(records)


def export_to_excel(transactions, filepath):
    """
    Export all data to Excel file with multiple sheets
    """
    wb = Workbook()

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    # Daily Summary Sheet
    daily_df = create_daily_summary(transactions)
    ws_daily = wb.active
    ws_daily.title = "Daily Summary"

    # Add headers
    headers = ['Date', 'Total Sales', 'Transaction Count', 'Items Count', 'Total VAT']
    for col_idx, header in enumerate(headers, 1):
        cell = ws_daily.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Add data
    for row_idx, row in enumerate(daily_df.values, 2):
        for col_idx, value in enumerate(row, 1):
            ws_daily.cell(row=row_idx, column=col_idx, value=value)

    # Adjust column widths
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws_daily.column_dimensions[col].width = 15

    # Detailed Transactions Sheet
    trans_df = create_detailed_transactions_df(transactions)
    ws_trans = wb.create_sheet("Transactions")

    headers = list(trans_df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws_trans.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, row in enumerate(trans_df.values, 2):
        for col_idx, value in enumerate(row, 1):
            ws_trans.cell(row=row_idx, column=col_idx, value=value)

    for i, col in enumerate(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']):
        ws_trans.column_dimensions[col].width = 12

    # Items Summary Sheet
    items_df = create_items_summary_df(transactions)
    ws_items = wb.create_sheet("Items Summary")

    headers = list(items_df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws_items.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, row in enumerate(items_df.values, 2):
        for col_idx, value in enumerate(row, 1):
            ws_items.cell(row=row_idx, column=col_idx, value=value)

    ws_items.column_dimensions['A'].width = 30
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws_items.column_dimensions[col].width = 15

    # Save
    wb.save(filepath)

    return daily_df, trans_df, items_df